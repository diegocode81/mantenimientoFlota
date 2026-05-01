#!/usr/bin/env python3
import argparse
import os
import re
import sys
import uuid
from datetime import datetime
from pathlib import Path

import psycopg2
from openpyxl import load_workbook


DEFAULT_EXCEL = (
    "/Users/diegomolina/Downloads/"
    "CONSOLIDADO DE FLOTA POWER BI TABLA VEHICULO.xlsx"
)
DATE_COLUMNS = [
    ("27/4/2026", "2026-04-27"),
    ("28/4/2026", "2026-04-28"),
    ("29/4/2026", "2026-04-29"),
    ("30/4/2026", "2026-04-30"),
    ("1/5/2026", "2026-05-01"),
]


def load_env(path: Path) -> None:
    if not path.exists():
        return

    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        value = value.strip().strip('"').strip("'")
        os.environ.setdefault(key.strip(), value)


def normalize_header(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).upper()


def normalize_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def parse_year(value):
    original = normalize_value(value)
    if not original:
        return None, None
    try:
        numeric = float(original)
    except ValueError:
        return None, original
    if numeric.is_integer() and 1900 <= numeric <= 2100:
        return int(numeric), original
    return None, original


def new_id(prefix):
    return f"{prefix}_{uuid.uuid4().hex}"


def read_rows(excel_path: Path):
    workbook = load_workbook(excel_path, data_only=True)
    worksheet = workbook["TABLA VEHICULOS "]
    headers = {
        normalize_header(worksheet.cell(1, column).value): column
        for column in range(1, worksheet.max_column + 1)
    }

    def cell(row, header):
        return worksheet.cell(row, headers[normalize_header(header)]).value

    vehicles = []
    for row in range(2, worksheet.max_row + 1):
        if not any(
            worksheet.cell(row, column).value not in (None, "")
            for column in range(1, worksheet.max_column + 1)
        ):
            continue

        year, year_original = parse_year(cell(row, "ANO"))
        vehicle = {
            "row": row,
            "itemOrigen": normalize_value(cell(row, "ITEM")),
            "placa": normalize_value(cell(row, "PLACA")),
            "disco": normalize_value(cell(row, "DISCO")),
            "marca": normalize_value(cell(row, "MARCA")),
            "tipo": normalize_value(cell(row, "TIPO")) or "",
            "ano": year,
            "anoOriginal": year_original,
            "cia": normalize_value(cell(row, "CIA")),
            "observaciones": normalize_value(cell(row, "OBSERVACIONES")),
            "rutaUbicacion": normalize_value(cell(row, "RUTA-UBICACIÓN")) or "",
            "tecnicosDesignados": normalize_value(cell(row, "TECNICOS DESIGNADOS"))
            or "",
            "statuses": {
                header: normalize_value(cell(row, header)) for header, _ in DATE_COLUMNS
            },
        }
        vehicles.append(vehicle)

    return vehicles


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=DEFAULT_EXCEL)
    parser.add_argument("--execute", action="store_true")
    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parents[1]
    load_env(repo_root / ".env")

    database_url = os.environ.get("DIRECT_URL") or os.environ.get("DATABASE_URL")
    if not database_url:
        print("DATABASE_URL or DIRECT_URL is required.", file=sys.stderr)
        return 1

    rows = read_rows(Path(args.excel))
    maintenance_count = len(rows) * len(DATE_COLUMNS)
    print(f"Vehicles to import: {len(rows)}")
    print(f"Maintenance records to import: {maintenance_count}")

    plates = [row["placa"] for row in rows if row["placa"]]
    with psycopg2.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute('SELECT COUNT(*) FROM "Vehiculo" WHERE "itemOrigen" IS NOT NULL')
            already_imported = cur.fetchone()[0]
            if already_imported:
                print(
                    f"Abort: {already_imported} imported vehicles already exist.",
                    file=sys.stderr,
                )
                return 2

            if plates:
                cur.execute(
                    'SELECT "placa" FROM "Vehiculo" WHERE "placa" = ANY(%s)',
                    (plates,),
                )
                conflicts = [record[0] for record in cur.fetchall()]
                if conflicts:
                    print(
                        "Abort: existing plates would conflict: "
                        + ", ".join(conflicts),
                        file=sys.stderr,
                    )
                    return 3

            if not args.execute:
                print("Dry run only. Re-run with --execute to import.")
                return 0

            now = datetime.utcnow()
            for row in rows:
                vehicle_id = new_id("veh")
                cur.execute(
                    """
                    INSERT INTO "Vehiculo"
                    ("id", "itemOrigen", "placa", "disco", "marca", "tipo",
                     "ano", "anoOriginal", "cia", "createdAt", "updatedAt")
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        vehicle_id,
                        row["itemOrigen"],
                        row["placa"],
                        row["disco"],
                        row["marca"],
                        row["tipo"],
                        row["ano"],
                        row["anoOriginal"],
                        row["cia"],
                        now,
                        now,
                    ),
                )

                for header, iso_date in DATE_COLUMNS:
                    status = row["statuses"][header]
                    if status not in {"OPERATIVO", "MANTENIMIENTO"}:
                        raise ValueError(
                            f"Invalid status at row {row['row']} / {header}: {status}"
                        )
                    cur.execute(
                        """
                        INSERT INTO "MantenimientoVehiculo"
                        ("id", "vehiculoId", "fechaMantenimiento", "estado",
                         "kilometrajeOdometro", "tipoMantenimiento",
                         "rutaUbicacion", "tecnicosDesignados", "observaciones",
                         "createdAt", "updatedAt")
                        VALUES (%s, %s, %s, %s::"EstadoVehiculo", %s,
                                %s::"TipoMantenimiento", %s, %s, %s, %s, %s)
                        """,
                        (
                            new_id("mnt"),
                            vehicle_id,
                            iso_date,
                            status,
                            0,
                            "PREVENTIVO",
                            row["rutaUbicacion"],
                            row["tecnicosDesignados"],
                            row["observaciones"],
                            now,
                            now,
                        ),
                    )
            conn.commit()

    print("Import completed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
